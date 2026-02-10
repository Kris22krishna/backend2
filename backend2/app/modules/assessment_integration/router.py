from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from app.db.session import get_db
from app.core.security import get_current_user
from app.modules.auth.models import User
from app.modules.assessment_integration.models import AssessmentStudent
from app.modules.assessment_integration.schemas import AssessmentStudentSchema, AssessmentAccessLogin
from app.core.security import create_access_token, oauth2_scheme, decode_token
from datetime import datetime
import openpyxl
from io import BytesIO
import random
import json
from uuid import UUID
from app.modules.questions.models import QuestionTemplate
from app.modules.questions.executor import executor
from app.modules.assessment_integration.models import AssessmentSession, AssessmentSessionQuestion, AssessmentStudent
from app.modules.assessment_integration.schemas import (
    AssessmentStudentSchema, AssessmentAccessLogin, 
    AssessmentSessionResponse, AssessmentQuestionResponse, AssessmentReport,
    AssessmentSubmission, AssessmentSessionDetail, AssessmentQuestionDetail
)

router = APIRouter(prefix="/assessment-integration", tags=["assessment-integration"])

def get_current_student(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    payload = decode_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Could not validate credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    student_id_str: str = payload.get("sub")
    if student_id_str is None:
        raise HTTPException(status_code=401, detail="Invalid token")
        
    try:
        student_id = UUID(student_id_str)
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid user ID")

    student = db.query(AssessmentStudent).filter(AssessmentStudent.id == student_id).first()
    if not student:
        raise HTTPException(status_code=401, detail="Student not found")
    return student

@router.post("/upload-students")
async def upload_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload Excel file with student details.
    Expected columns: Serial Number, Student Name, Grade, School Name, Phone Number
    """
    # Verify user is uploader or admin
    if current_user.user_type not in ["uploader", "admin", "assessment_uploader"]:
         raise HTTPException(status_code=403, detail="Access denied")

    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        
        # Iterate rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
             raise HTTPException(status_code=400, detail="Empty file")
             
        # Extract headers (assuming first row)
        headers = [str(h).lower().strip() if h else '' for h in rows[0]]
        
        # Map columns
        try:
            idx_serial = headers.index('serial number')
            idx_name = headers.index('student name')
            idx_grade = headers.index('grade')
            idx_school = headers.index('school name')
            idx_phone = headers.index('phone number')
        except ValueError as e:
             raise HTTPException(status_code=400, detail=f"Missing required column: {str(e)}")

        uploaded_count = 0
        errors = []
        
        for i, row in enumerate(rows[1:], start=2): # Start from row 2
            try:
                serial = str(row[idx_serial]).strip() if row[idx_serial] else None
                name = str(row[idx_name]).strip() if row[idx_name] else None
                grade = str(row[idx_grade]).strip() if row[idx_grade] else None
                school = str(row[idx_school]).strip() if row[idx_school] else None
                phone = str(row[idx_phone]).strip() if row[idx_phone] else None
                
                if not serial or not name or not grade or not school:
                    continue # Skip empty rows
                
                # Check for duplicate serial number
                existing = db.query(AssessmentStudent).filter(AssessmentStudent.serial_number == serial).first()
                if existing:
                    # Update or Skip? Let's Skip for now and report
                    errors.append(f"Row {i}: Serial Number {serial} already exists.")
                    continue
                
                new_student = AssessmentStudent(
                    serial_number=serial,
                    name=name,
                    grade=grade,
                    school_name=school,
                    phone_number=phone,
                    uploaded_by_user_id=current_user.user_id
                )
                db.add(new_student)
                uploaded_count += 1
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
        
        db.commit()
        
        return {
            "success": True,
            "data": {
                "uploaded_count": uploaded_count,
                "errors": errors
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

@router.get("/uploaded-students", response_model=list[AssessmentStudentSchema])
def get_uploaded_students(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    List students uploaded by the current user.
    """
    students = db.query(AssessmentStudent).filter(
        AssessmentStudent.uploaded_by_user_id == current_user.user_id
    ).all()
    return students

@router.get("/dashboard-stats")
def get_dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get statistics for the uploader dashboard.
    """
    # 1. Total Students
    student_query = db.query(AssessmentStudent).filter(
        AssessmentStudent.uploaded_by_user_id == current_user.user_id
    )
    total_students = student_query.count()
    student_ids = [s.id for s in student_query.all()]

    # 2. Total Assessments
    assessments_count = db.query(AssessmentSession).filter(
        AssessmentSession.student_id.in_(student_ids)
    ).count()

    # 3. High Scorers (Accuracy > 80% in at least one COMPLETED session)
    # We define accuracy as (count of is_correct == 'True') / (total questions in session)
    
    high_scorers_count = 0
    if student_ids:
        # Fetch all completed sessions for these students
        # OPTIMIZED: Use selectinload to eagerly load questions (prevents N+1)
        from sqlalchemy.orm import selectinload
        
        sessions = db.query(AssessmentSession).options(
            selectinload(AssessmentSession.questions)
        ).filter(
            AssessmentSession.student_id.in_(student_ids),
            AssessmentSession.status == "COMPLETED"
        ).all()

        high_scorer_student_ids = set()
        for session in sessions:
            questions = session.questions
            if not questions:
                continue
            
            correct_count = sum(1 for q in questions if q.is_correct == 'True')
            accuracy = correct_count / len(questions)
            
            if accuracy >= 0.8:
                high_scorer_student_ids.add(session.student_id)
        
        high_scorers_count = len(high_scorer_student_ids)


    return {
        "success": True,
        "data": {
            "total_students": total_students,
            "assessments_count": assessments_count,
            "high_scorers_count": high_scorers_count,
            "top_grade": _get_top_grade(student_query.all())
        }
    }

def _get_top_grade(students):
    if not students:
        return "N/A"
    
    from collections import Counter
    import re
    
    grades = []
    for s in students:
        match = re.search(r'\d+', str(s.grade))
        if match:
            grades.append(match.group())
    
    if not grades:
        return "N/A"
        
    most_common = Counter(grades).most_common(1)
    return most_common[0][0] if most_common else "N/A"

@router.get("/reports", response_model=list[AssessmentReport])
def get_reports(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get completed assessment reports for students uploaded by the current uploader.
    """
    # 1. Fetch student IDs for this uploader
    student_ids = db.query(AssessmentStudent.id).filter(
        AssessmentStudent.uploaded_by_user_id == current_user.user_id
    ).all()
    student_ids = [s[0] for s in student_ids]

    if not student_ids:
        return []

    # 2. Fetch completed sessions for these students
    # OPTIMIZED: Use eager loading to prevent N+1 queries
    from sqlalchemy.orm import joinedload, selectinload
    
    reports = db.query(AssessmentSession).options(
        joinedload(AssessmentSession.student),
        selectinload(AssessmentSession.questions)
    ).filter(
        AssessmentSession.student_id.in_(student_ids),
        AssessmentSession.status == "COMPLETED"
    ).order_by(AssessmentSession.completed_at.desc()).all()

    # 3. Format the data
    result = []
    for report in reports:
        correct_count = sum(1 for q in report.questions if q.is_correct == 'True')
        total_questions = len(report.questions)
        accuracy = (correct_count / total_questions * 100) if total_questions > 0 else 0
        
        result.append({
            "id": report.id,
            "student_id": report.student_id,
            "student_name": report.student.name,
            "grade": report.student.grade,
            "completed_at": report.completed_at,
            "total_questions": total_questions,
            "correct_answers": correct_count,
            "accuracy": round(accuracy, 2)
        })
    
    return result


@router.get("/reports/export")
def export_reports(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export completed assessment reports for students uploaded by the current uploader to Excel.
    """
    # 1. Fetch student IDs for this uploader
    student_ids = db.query(AssessmentStudent.id).filter(
        AssessmentStudent.uploaded_by_user_id == current_user.user_id
    ).all()
    student_ids = [s[0] for s in student_ids]

    if not student_ids:
        raise HTTPException(status_code=404, detail="No students found to export reports for.")

    # 2. Fetch completed sessions for these students
    reports = db.query(AssessmentSession).filter(
        AssessmentSession.student_id.in_(student_ids),
        AssessmentSession.status == "COMPLETED"
    ).order_by(AssessmentSession.completed_at.desc()).all()

    if not reports:
        raise HTTPException(status_code=404, detail="No completed assessment reports found to export.")

    # 3. Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Reports"

    # Headers
    headers = ["Serial Number", "Student Name", "Grade", "School Name", "Completion Date", "Score", "Total Questions", "Accuracy (%)"]
    ws.append(headers)

    # Data
    for report in reports:
        correct_count = sum(1 for q in report.questions if q.is_correct == 'True')
        total_questions = len(report.questions)
        accuracy = (correct_count / total_questions * 100) if total_questions > 0 else 0
        
        ws.append([
            report.student.serial_number,
            report.student.name,
            report.student.grade,
            report.student.school_name,
            report.completed_at.strftime("%Y-%m-%d %H:%M:%S") if report.completed_at else "N/A",
            correct_count,
            total_questions,
            round(accuracy, 2)
        ])

    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Assessment_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/reports/{session_id}", response_model=AssessmentSessionDetail)
def get_report_detail(
    session_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get detailed report for a specific assessment session.
    """
    # 1. Fetch the session
    session = db.query(AssessmentSession).filter(
        AssessmentSession.id == session_id
    ).first()
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
        
    # 2. Verify access (student must belong to uploader)
    if session.student.uploaded_by_user_id != current_user.user_id:
         raise HTTPException(status_code=403, detail="Access denied")
         
    # 3. Compile Details
    questions_data = []
    correct_count = 0
    total_questions = len(session.questions)
    
    for q in session.questions:
        is_correct = q.is_correct == 'True'
        if is_correct:
            correct_count += 1
            status_str = "correct"
        elif not q.student_answer:
            status_str = "skipped"
        else:
            status_str = "wrong"
            
        questions_data.append({
            "id": q.id,
            "question_html": q.question_html,
            "student_answer": q.student_answer,
            "correct_answer": q.correct_answer,
            "is_correct": is_correct,
            "status": status_str
        })
        
    accuracy = (correct_count / total_questions * 100) if total_questions > 0 else 0
    
    duration = 0
    if session.completed_at and session.started_at:
        duration = (session.completed_at - session.started_at).total_seconds() / 60.0
        
    return {
        "session_id": session.id,
        "student_name": session.student.name,
        "grade": session.student.grade,
        "started_at": session.started_at,
        "completed_at": session.completed_at,
        "duration_minutes": round(duration, 2),
        "score": correct_count,
        "total_questions": total_questions,
        "accuracy": round(accuracy, 2),
        "questions": questions_data
    }

@router.post("/student-access")
def student_access(
    login_in: AssessmentAccessLogin,
    db: Session = Depends(get_db)
):
    """
    Login using Serial Number. Returns a temporary token.
    """
    student = db.query(AssessmentStudent).filter(
        AssessmentStudent.serial_number == login_in.serial_number
    ).first()
    
    if not student:
        raise HTTPException(status_code=401, detail="Invalid Serial Number")
        
    # Generate a token with student specifics
    # We might need a generic user_id or a special type in token
    # For now, let's create a token with a special subject prefix or just the ID
    
    token = create_access_token(
        user_id=str(student.id),
        tenant_id=str(student.uploader.tenant_id)
    )
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "student_name": student.name,
        "grade": student.grade,
        "school_name": student.school_name,
        "serial_number": student.serial_number
    }

@router.post("/start-assessment", response_model=AssessmentSessionResponse)
def start_assessment(
    student: AssessmentStudent = Depends(get_current_student),
    db: Session = Depends(get_db)
):
    """
    Start a new assessment session for the student.
    Generates 25 'Hard' questions for the student's grade.
    """
    
    import random

    # 1. Check for existing active session
    active_session = db.query(AssessmentSession).filter(
        AssessmentSession.student_id == student.id,
        AssessmentSession.status.in_(["PENDING", "IN_PROGRESS"])
    ).first()
    
    if active_session:
        # Fetch questions with topic from template (V1/V2)
        # OPTIMIZED: Bulk fetch all templates to avoid N+1 query problem
        
        questions = db.query(AssessmentSessionQuestion).filter(
            AssessmentSessionQuestion.session_id == active_session.id
        ).all()
        
        # Collect all template IDs
        template_ids = [q.template_id for q in questions]
        
        if template_ids:
            # Bulk fetch V2 templates (likely most questions)
            from app.modules.questions.models import QuestionGeneration
            v2_templates = {
                t.template_id: t 
                for t in db.query(QuestionGeneration).filter(
                    QuestionGeneration.template_id.in_(template_ids)
                ).all()
            }
            
            # Bulk fetch V1 templates (for any not in V2)
            v1_templates = {
                t.template_id: t
                for t in db.query(QuestionTemplate).filter(
                    QuestionTemplate.template_id.in_(template_ids)
                ).all()
            }
            
            # Map topics to questions (in-memory, no extra queries)
            for q in questions:
                if q.template_id in v2_templates:
                    q.topic = v2_templates[q.template_id].skill_name
                elif q.template_id in v1_templates:
                    q.topic = v1_templates[q.template_id].topic
                else:
                    q.topic = "Unknown"  # Fallback
        
        return {
            "session_id": active_session.id,
            "questions": questions,
            "duration_minutes": 30
        }
    
    # 2. Extract Grade
    import re
    match = re.search(r'\d+', str(student.grade))
    grade_level = int(match.group()) if match else None
    
    if not grade_level:
         raise HTTPException(status_code=400, detail="Could not determine valid grade from student record")
         
    # 3. Define Selection Logic
    selected_items = []
    
    # SPECIAL HANDLING FOR GRADE 7
    if grade_level == 7:
        from app.modules.assessment_integration.grade7_blueprint import GRADE_7_TEMPLATE_IDS
        from app.modules.questions.models import QuestionGeneration

        # Fetch the specific 25 templates
        templates = db.query(QuestionGeneration).filter(
            QuestionGeneration.template_id.in_(GRADE_7_TEMPLATE_IDS)
        ).all()
        
        # Verify we have all of them (or as many as possible)
        if not templates:
             raise HTTPException(status_code=404, detail="Grade 7 assessment templates not found")
             
        # Normalize to standard format
        for t in templates:
            selected_items.append({
                "id": t.template_id,
                "source": "v2",
                "topic": t.skill_name,
                "difficulty": t.difficulty,
                "code": t.question_template,
                "answer_code": t.answer_template,
                "type": t.type,
                "obj": t
            })
            
        # Shuffle to randomize order
        random.shuffle(selected_items)
        
    else:
        # STANDARD LOGIC FOR OTHER GRADES
        if grade_level <= 5:
            target_mcq = 18
            target_input = 7
        elif grade_level <= 8:
            target_mcq = 13
            target_input = 12
        else:
            target_mcq = 10
            target_input = 15
            
        total_target = target_mcq + target_input
        
        # 4. Fetch Templates (Strictly Easy)
        from app.modules.questions.models import QuestionGeneration
        
        # Fetch Easy V2 Templates
        v2_easy = db.query(QuestionGeneration).filter(
            QuestionGeneration.grade == grade_level,
            QuestionGeneration.difficulty.ilike('%easy%')
        ).all()
        
        # Fetch Easy V1 Templates
        v1_easy = db.query(QuestionTemplate).filter(
            QuestionTemplate.status == "active",
            QuestionTemplate.grade_level.any(grade_level),
            QuestionTemplate.difficulty.ilike('%easy%')
        ).all()
        
        def normalize_v2(t):
            return {
                "id": t.template_id,
                "source": "v2",
                "topic": t.skill_name,
                "difficulty": t.difficulty,
                "code": t.question_template,
                "answer_code": t.answer_template,
                "type": t.type,
                "obj": t
            }
            
        def normalize_v1(t):
            return {
                "id": t.template_id,
                "source": "v1",
                "topic": t.topic,
                "difficulty": t.difficulty,
                "code": t.dynamic_question,
                "type": t.type,
                "obj": t
            }
    
        # Categorize pools by type
        mcq_pool = []
        input_pool = []
        
        for t in v2_easy:
            norm = normalize_v2(t)
            q_type = norm['type'].lower() if norm['type'] else ""
            if q_type == 'mcq':
                mcq_pool.append(norm)
            else:
                input_pool.append(norm)
                
        for t in v1_easy:
            norm = normalize_v1(t)
            q_type = norm['type'].lower() if norm['type'] else ""
            if q_type == 'mcq':
                mcq_pool.append(norm)
            else:
                input_pool.append(norm)
    
        # Fallback: if easy pool is too small, we might need to take other difficulties 
        # but the user said "use the easy templates already there", implying they exist.
        # To be safe, if we are short, we'll search for others but warn.
        if len(mcq_pool) < target_mcq or len(input_pool) < target_input:
            print(f"DEBUG: Shortage in Easy templates. MCQ: {len(mcq_pool)}/{target_mcq}, Input: {len(input_pool)}/{target_input}")
            # Fallback to medium if really needed to maintain the 25 count
            if len(mcq_pool) < target_mcq:
                v2_med_mcq = db.query(QuestionGeneration).filter(
                    QuestionGeneration.grade == grade_level,
                    QuestionGeneration.difficulty.ilike('%medium%'),
                    QuestionGeneration.type == 'mcq'
                ).all()
                mcq_pool.extend([normalize_v2(t) for t in v2_med_mcq])
                
            if len(input_pool) < target_input:
                v2_med_input = db.query(QuestionGeneration).filter(
                    QuestionGeneration.grade == grade_level,
                    QuestionGeneration.difficulty.ilike('%medium%'),
                    QuestionGeneration.type != 'mcq'
                ).all()
                input_pool.extend([normalize_v2(t) for t in v2_med_input])
    
        if not mcq_pool and not input_pool:
            raise HTTPException(status_code=404, detail="No assessment questions available for this grade level")
    
        # 5. Select items while maintaining topic diversity
        def select_from_pool(pool, count):
            if not pool: return []
            if len(pool) <= count: return pool
            
            # Group by topic
            topic_map = {}
            for item in pool:
                t = item['topic']
                if t not in topic_map:
                    topic_map[t] = []
                topic_map[t].append(item)
                
            selected = []
            topics = list(topic_map.keys())
            random.shuffle(topics)
            
            # Round 1: Unique topics
            for t in topics:
                if len(selected) < count:
                    selected.append(random.choice(topic_map[t]))
                    
            # Round 2: Fill rest
            while len(selected) < count:
                t = random.choice(topics)
                # Find an item not already selected if possible
                available = [i for i in topic_map[t] if i not in selected]
                if available:
                    selected.append(random.choice(available))
                else:
                    selected.append(random.choice(topic_map[t]))
            
            return selected
    
        selected_items.extend(select_from_pool(mcq_pool, target_mcq))
        selected_items.extend(select_from_pool(input_pool, target_input))
        
        # Last check if we are still short due to empty pools
        if len(selected_items) < total_target:
            # Fill from whichever pool has more
            remaining = total_target - len(selected_items)
            # (Actually, we just take anything left from both)
            extra_candidates = [i for i in (mcq_pool + input_pool) if i not in selected_items]
            if extra_candidates:
                random.shuffle(extra_candidates)
                selected_items.extend(extra_candidates[:remaining])

    # Create Session
    session = AssessmentSession(
        student_id=student.id,
        status="PENDING",
        started_at=datetime.utcnow()
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    
    generated_questions = []
    questions_to_insert = []  # For bulk insert
            
    for idx, template_data in enumerate(selected_items):
        try:
            result = {}
            
            # Execute based on source type
            if template_data['source'] == 'v2':
                q_code = template_data['code']
                a_code = template_data.get('answer_code', '')
                
                # Cleanup markdown
                if q_code.startswith("```python"): q_code = q_code.replace("```python", "").replace("```", "")
                if a_code.startswith("```python"): a_code = a_code.replace("```python", "").replace("```", "")
                
                # Use sequential execution for V2 (shared context)
                result = executor.execute_sequential([q_code, a_code])
                
            else: # v1
                code = template_data['code']
                if code.startswith("```python"): code = code.replace("```python", "").replace("```", "")
                result = executor.execute_generator(code)
            
            question_text = result.get('question', '')
            answer_value = str(result.get('answer', ''))
            
            # Type handling (ensure lowercase for DB consistency if needed)
            q_type = result.get('type')
            if q_type: q_type = q_type.lower()
            
            # Infer MCQ if options exist
            if 'options' in result and result['options'] and isinstance(result['options'], list):
                if not q_type or q_type == 'user_input':
                    q_type = 'mcq'
            
            # Default if still nothing
            if not q_type:
                orig_type = template_data['type']
                q_type = orig_type.lower() if orig_type else 'user_input'
            
            options = json.dumps(result.get('options', [])) if 'options' in result else None
            
            # Prepare question data for bulk insert
            question_data = {
                'session_id': session.id,
                'template_id': template_data['id'],
                'question_html': question_text,
                'question_type': q_type,
                'options': options,
                'correct_answer': answer_value,
                'student_answer': None,
                'is_correct': None
            }
            questions_to_insert.append(question_data)
            
            # Create a mock object for the response (with topic attached)
            # We'll need to retrieve these from DB after bulk insert
            mock_q = type('obj', (object,), {
                **question_data,
                'topic': template_data['topic'],
                'id': None  # Will be assigned after insert
            })()
            generated_questions.append(mock_q)
            
        except Exception as e:
            print(f"Failed to generate question from template {template_data['id']} ({template_data['source']}): {e}")
            continue
    
    # OPTIMIZED: Bulk insert all questions at once instead of individual db.add()
    if questions_to_insert:
        db.bulk_insert_mappings(AssessmentSessionQuestion, questions_to_insert)
    
    db.commit()
    
    # Fetch the actually inserted questions with their IDs and attach topics
    actual_questions = db.query(AssessmentSessionQuestion).filter(
        AssessmentSessionQuestion.session_id == session.id
    ).all()
    
    # Map topics to the actual questions
    template_id_to_topic = {gq.template_id: gq.topic for gq in generated_questions}
    for q in actual_questions:
        q.topic = template_id_to_topic.get(q.template_id, "Unknown")
    
    
    return {
        "session_id": session.id,
        "questions": actual_questions,
        "duration_minutes": 30
    }


@router.post("/submit-assessment")
def submit_assessment(
    submission: AssessmentSubmission,
    student: AssessmentStudent = Depends(get_current_student),
    db: Session = Depends(get_db)
):
    """
    Submit assessment answers. 
    1. Updates student answers in AssessmentSessionQuestion
    2. Auto-grading: Checks answers against correct_answer
    3. Marks Session as COMPLETED
    4. Records completed_at
    """
    
    
    # Check session
    session = db.query(AssessmentSession).filter(
        AssessmentSession.id == submission.session_id,
        AssessmentSession.student_id == student.id
    ).first()
    
    if not session:
        raise HTTPException(status_code=404, detail="Active session not found or already completed")
    
    if session.status not in ["PENDING", "IN_PROGRESS"]:
        raise HTTPException(status_code=404, detail=f"Active session not found. Current status: {session.status}")
        
    # Process Answers
    questions = db.query(AssessmentSessionQuestion).filter(
        AssessmentSessionQuestion.session_id == session.id
    ).all()
    
    for q in questions:
        q_id_str = str(q.id)
        if q_id_str in submission.answers:
            student_a = submission.answers[q_id_str]
            q.student_answer = student_a
            
            # Auto-grade
            # Simple exact match for now. In future, we can use LLM grading or smarter comparison.
            # Handle float comparison for numbers?
            
            is_correct = False
            
            correct = str(q.correct_answer).strip().lower()
            given = str(student_a).strip().lower()
            
            if correct == given:
                is_correct = True
            else:
                 # Try removing potential Markdown wrappers if exact match fails
                 # e.g. "8.44" vs "**8.44**"
                 import re
                 c_clean = re.sub(r'[*_`]', '', correct)
                 g_clean = re.sub(r'[*_`]', '', given)
                 if c_clean == g_clean:
                     is_correct = True
            
            q.is_correct = str(is_correct) # API expects string 'True'/'False' for now per legacy logic
            
    session.status = "COMPLETED"
    session.completed_at = datetime.utcnow()
    db.commit()
    
    return {"status": "success", "message": "Assessment submitted successfully"}
